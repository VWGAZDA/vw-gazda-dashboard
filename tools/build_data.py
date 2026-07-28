#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aktualizacja dashboardu GAZDA Katowice z RAPORT_ZBIORCZY_*.xlsx

Uzycie:
  python3 build_data.py --xlsx <plik.xlsx> --html <index.html> [--out <index.html>]

Sekcje odswiezane z raportu VW:
  meta, realizacja, cele (wykonanie), stan_aek, handlowcy_detal, handlowcy_flota,
  kanaly, magazyn, przet, dostawy, prognoza, prognoza_handlowcy, prognoza_mies,
  ranking_combined

Sekcje zachowywane z istniejacego index.html (dane wprowadzane recznie):
  cele.cel_zespolowy / cele[].cel / bez_celu / zadania, kpi_handlowcy, kpi_cele,
  koniunktura, opinie_google, trade_in_crm, produkcja, push, grafik,
  filtered_records_count, filtered_clients
"""
import argparse, datetime as dt, json, re, sys, unicodedata
from collections import defaultdict, OrderedDict

import openpyxl

MIESIACE_PL = ['stycznia', 'lutego', 'marca', 'kwietnia', 'maja', 'czerwca',
               'lipca', 'sierpnia', 'września', 'października', 'listopada', 'grudnia']
MIESIACE_MIAN = ['styczeń', 'luty', 'marzec', 'kwiecień', 'maj', 'czerwiec',
                 'lipiec', 'sierpień', 'wrzesień', 'październik', 'listopad', 'grudzień']


# ---------------------------------------------------------------- helpers

def norm(s):
    """Klucz porownawczy imienia/nazwiska: bez ogonkow, uppercase, posortowane tokeny."""
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = s.replace('ł', 'l').replace('Ł', 'L')
    s = re.sub(r'[^A-Za-z\- ]', ' ', s).upper()
    return ' '.join(sorted(t for t in re.split(r'[\s\-]+', s) if t))


def pretty_name(s):
    """'PAWLAS-POLOCZEK MALGORZATA' -> 'Malgorzata Pawlas-Poloczek'"""
    if not s:
        return ''
    parts = [p for p in str(s).split() if p]
    if len(parts) < 2:
        return str(s).title()
    first = parts[-1]
    rest = parts[:-1]
    cap = lambda w: '-'.join(x.capitalize() for x in w.split('-'))
    return cap(first) + ' ' + ' '.join(cap(w) for w in rest)


def d2s(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str) and re.match(r'^\d{4}-\d{2}-\d{2}', v):
        return v[:10]
    return ''


def d2d(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str) and re.match(r'^\d{4}-\d{2}-\d{2}', v):
        return dt.date(int(v[:4]), int(v[5:7]), int(v[8:10]))
    return None


def num(v, default=0):
    if v is None or v == '' or v == '-':
        return default
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(',', '.'))
    except ValueError:
        return default


def i(v, default=0):
    return int(round(num(v, default)))


class Sheet:
    """Arkusz bazowy: naglowek w wierszu 0, dostep przez nazwe kolumny."""

    def __init__(self, wb, name):
        self.name = name
        self.rows = []
        if name not in wb.sheetnames:
            self.cols = {}
            return
        raw = list(wb[name].iter_rows(values_only=True))
        if not raw:
            self.cols = {}
            return
        hdr = ['' if c is None else str(c).strip() for c in raw[0]]
        self.cols = {h: k for k, h in enumerate(hdr) if h}
        for r in raw[1:]:
            if r is None or all(c is None or c == '' for c in r):
                continue
            self.rows.append(r)

    def g(self, row, col, default=None):
        k = self.cols.get(col)
        if k is None or k >= len(row):
            return default
        v = row[k]
        return default if v is None else v

    def s(self, row, col):
        v = self.g(row, col, '')
        return '' if v is None else str(v).strip()

    def __iter__(self):
        return iter(self.rows)

    def __len__(self):
        return len(self.rows)


def kv_sheet(wb, name):
    """Arkusze REALIZACJA*: kolumna A = etykieta, kolumna B = wartosc."""
    out = {}
    if name not in wb.sheetnames:
        return out
    for r in wb[name].iter_rows(values_only=True):
        if not r or r[0] is None:
            continue
        key = str(r[0]).strip()
        if key and key not in out:
            out[key] = r[1] if len(r) > 1 else None
    return out


# ---------------------------------------------------------------- extract DATA z html

DATA_RE = re.compile(r'const\s+DATA\s*=\s*')


def find_data_span(html):
    m = DATA_RE.search(html)
    if not m:
        raise SystemExit('BLAD: nie znaleziono "const DATA =" w pliku HTML')
    start = html.index('{', m.end() - 1)
    depth, in_str, esc, quote = 0, False, False, ''
    for k in range(start, len(html)):
        ch = html[k]
        if in_str:
            if esc:
                esc = False
            elif ch == '\\':
                esc = True
            elif ch == quote:
                in_str = False
            continue
        if ch in '"\'':
            in_str, quote = True, ch
        elif ch == '{':
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0:
                return m.start(), start, k + 1
    raise SystemExit('BLAD: niezamkniety obiekt DATA')


# ---------------------------------------------------------------- build

def apply_manual(new, man, rap_d, roster, flota_keys):
    """Nadpisuje sekcje pochodzace z arkuszy Google / wpisow recznych."""
    if man.get('ih_numer'):
        new['meta']['ih_numer'] = man['ih_numer']
    if man.get('zrodla_dodatkowe'):
        new['meta']['zrodla_dodatkowe'] = man['zrodla_dodatkowe']
    if man.get('cel_zespolowy'):
        c = new['cele']
        c['cel_zespolowy'] = i(man['cel_zespolowy'])
        c['brakuje'] = max(0, c['cel_zespolowy'] - c['wykonanie_realne'])
        c['pct'] = int(round(c['wykonanie_realne'] / c['cel_zespolowy'] * 100)) if c['cel_zespolowy'] else 0
    if man.get('bez_celu') is not None:
        new['cele']['bez_celu'] = man['bez_celu']

    # --- KPI handlowcy -------------------------------------------------
    # trade-in bierzemy z CRM (to samo zrodlo co zakladka Trade-in CRM)
    crm_leady = {}
    _t = man.get('trade_in_crm') or {}
    for row in ((_t.get('biezacy') or {}).get('ranking') or []):
        crm_leady[norm(row.get('name'))] = i(row.get('przekazane'))

    kpi = man.get('kpi')
    if kpi:
        name_by_key = {k: n for k, n, s, c in roster}
        rows = []
        for h in kpi:
            k = norm(h.get('name'))
            rows.append({
                'name': name_by_key.get(k, h.get('name')),
                'oferty': i(h.get('oferty')), 'klienci': i(h.get('klienci')),
                'zamowienia': i(h.get('zamowienia')), 'wydania': i(h.get('wydania')),
                'traffic': i(h.get('traffic')), 'opinie': i(h.get('opinie')),
                'trade_in': crm_leady.get(k, 0) if crm_leady else i(h.get('trade_in')),
                'flota': bool(h.get('flota', k in flota_keys)),
            })
        new['kpi_handlowcy'] = rows
        # zadania (Trade-in / Opinie Google) licza sie z KPI
        zad_def = man.get('zadania_cele') or {'Trade-in': 10, 'Opinie Google': 5}
        zadania = []
        for nazwa, pole in (('Trade-in', 'trade_in'), ('Opinie Google', 'opinie')):
            zadania.append({
                'nazwa': nazwa,
                'cel_ind': i(zad_def.get(nazwa, 10)),
                'handlowcy': [{'name': r['name'], 'cel': i(zad_def.get(nazwa, 10)),
                               'wykonanie': r[pole]} for r in rows if not r['flota']],
            })
        new['cele']['zadania'] = zadania
        n_detal = sum(1 for r in rows if not r['flota']) or 1
        base = man.get('kpi_cele_bazowe') or {}
        if base:
            new['kpi_cele'] = {k: i(v) * n_detal for k, v in base.items()}
    if man.get('kpi_cele'):
        new['kpi_cele'] = {k: i(v) for k, v in man['kpi_cele'].items()}

    # --- koniunktura ---------------------------------------------------
    kd = man.get('koniunktura_daily')
    if kd:
        dni = (dt.date(rap_d.year + (rap_d.month // 12), rap_d.month % 12 + 1, 1)
               - dt.timedelta(days=1)).day
        get = lambda key, d: i((kd.get(key) or [0] * 31)[d - 1]) if d - 1 < len(kd.get(key) or []) else 0
        raw = {d: get('oferty', d) + get('klienci', d) + get('traffic', d) for d in range(1, dni + 1)}
        wknd = {d: dt.date(rap_d.year, rap_d.month, d).weekday() >= 5 for d in range(1, dni + 1)}
        mx = max([raw[d] for d in raw if not wknd[d]] or [0]) or 1
        new['koniunktura'] = [{'d': d, 'raw': raw[d], 'v': round(min(raw[d] / mx, 1.0), 3),
                               'wknd': wknd[d]} for d in range(1, dni + 1)]

    # --- opinie Google -------------------------------------------------
    if man.get('opinie') is not None:
        prev = man.get('opinie_poprzedni_raport') or new.get('opinie_poprzedni_raport') or ''
        pd_ = None
        if re.match(r'^\d{2}\.\d{2}\.\d{4}$', prev):
            pd_ = dt.date(int(prev[6:]), int(prev[3:5]), int(prev[:2]))
        out = []
        for o in man['opinie']:
            ds = o.get('data', '')
            od = None
            if re.match(r'^\d{2}\.\d{2}\.\d{4}$', ds):
                od = dt.date(int(ds[6:]), int(ds[3:5]), int(ds[:2]))
            out.append({'data': ds, 'nazwa': o.get('nazwa', ''),
                        'handlowiec': o.get('handlowiec', ''),
                        'nowa': bool(od and pd_ and od > pd_),
                        'tresc': o.get('tresc', '')})
        out.sort(key=lambda x: (x['data'][6:], x['data'][3:5], x['data'][:2]), reverse=True)
        new['opinie_google'] = out
        new['opinie_poprzedni_raport'] = prev

    # --- trade-in CRM --------------------------------------------------
    t = man.get('trade_in_crm')
    if t:
        cur = t.get('biezacy') or {}
        prv = t.get('poprzedni') or {}
        new['trade_in_crm'] = {
            'miesiac': t.get('miesiac') or (rap_d.strftime('%Y-%m') if rap_d else ''),
            'miesiac_label': cur.get('label', ''),
            'aktywny': bool(cur.get('ranking') or cur.get('wygrane')),
            'biezacy': {'label': cur.get('label', ''), 'wygrane': cur.get('wygrane', []),
                        'ranking': cur.get('ranking', []), 'total_leads': i(cur.get('total_leads'))},
            'poprzedni': {'label': prv.get('label', ''), 'wygrane': prv.get('wygrane', []),
                          'ranking': prv.get('ranking', []), 'total_leads': i(prv.get('total_leads'))},
        }

    # --- informacja produkcyjna ---------------------------------------
    if man.get('produkcja'):
        new['produkcja'] = man['produkcja']
    if man.get('push') is not None:
        new['push'] = man['push']
    if man.get('grafik'):
        new['grafik'] = man['grafik']


def build(xlsx_path, old, man=None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    R = kv_sheet(wb, 'REALIZACJA')
    RD = kv_sheet(wb, 'REALIZACJA_detal')
    RF = kv_sheet(wb, 'REALIZACJA_flota')

    b_aak = Sheet(wb, 'baza_AAK')
    b_aek = Sheet(wb, 'baza_AEK')
    b_mag = Sheet(wb, 'baza_MAG_DE')
    b_90 = Sheet(wb, 'baza_age_rtt_90DNIbm')
    b_120 = Sheet(wb, 'baza_age_rtt_120DNIbm')
    b_wd = Sheet(wb, 'baza_WARTOŚĆ_DOSTAW')
    b_pl = Sheet(wb, 'baza_PIPE_LBOR')

    data_rap = d2s(R.get('data')) or d2s(RD.get('data'))
    rap_d = d2d(data_rap)
    uplyw = round(num(R.get('upływ_czasu')), 3)

    new = json.loads(json.dumps(old))  # deep copy

    # ---- meta -------------------------------------------------------
    mies_lbl = MIESIACE_PL[rap_d.month - 1] if rap_d else ''
    mies_mian = MIESIACE_MIAN[rap_d.month - 1] if rap_d else ''
    kvps = str(R.get('KVPS') or RD.get('KVPS') or '01380V')
    new['meta'] = dict(old.get('meta', {}))
    new['meta'].update({
        'dealer': 'GAZDA Katowice',
        'marka': 'Volkswagen',
        'miasto': 'Katowice',
        'kvps': kvps,
        'data_raportu': data_rap,
        'data_raportu_pl': '%d %s %d' % (rap_d.day, mies_lbl, rap_d.year) if rap_d else '',
        'zrodlo': 'RAPORT_ZBIORCZY %s, stan na %s, GAZDA KATOWICE %s' % (mies_mian, data_rap, kvps),
    })

    # ---- realizacja -------------------------------------------------
    def blok(K, detal):
        aak = i(K.get('AAK'))
        out = {
            'aak': aak,
            'aek': i(K.get('AEK')),
            'abk_bm': i(K.get('ABK_bm')),
            'fk': i(K.get('FK_bez_AAK')),
            'uplyw': uplyw,
        }
        if detal:
            out['aakmax'] = i(K.get('AAKmax'))
            out['prognoza'] = int(round(aak / uplyw)) if uplyw else 0
        return out

    new['realizacja'] = {
        'detal': blok(RD, True),
        'flota': blok(RF, False),
        'razem': {
            'aak': i(R.get('AAK')),
            'aek': i(R.get('AEK')),
            'abk_bm': i(R.get('ABK_bm')),
            'fk': i(R.get('FK_bez_AAK')),
            'aakmax': i(R.get('AAKmax')),
            'aeh': i(R.get('AEH')),
            'd_aeh': i(R.get('∆_AEH')),
            's180bm': i(R.get('180bm')),
            'mag_de': i(R.get('MAG_DE')),
            'kswbm': i(R.get('KSWbm')),
        },
    }
    # detal: aakmax musi byc >= aak
    new['realizacja']['detal']['aakmax'] = max(new['realizacja']['detal']['aakmax'],
                                               new['realizacja']['detal']['aak'])

    # ---- AAK per handlowiec, z podzialem detal / flota --------------
    # DETAL = true retail + small fleet + grupy zawodowe/others + funkcyjne
    # FLOTA = strategic fleet (medium/big fleet), obszar FLOTOWY
    aak_detal = defaultdict(int)
    aak_flota = defaultdict(int)
    aak_names = {}
    for r in b_aak:
        who = b_aak.s(r, 'odpowiedzialny')
        if not who:
            continue
        k = norm(who)
        aak_names.setdefault(k, pretty_name(who))
        obszar = b_aak.s(r, 'obszar').upper()
        kanal = b_aak.s(r, 'kanał_rodzaj_użytkownik').lower()
        funkcyjny = b_aak.s(r, 'status').upper() == 'FUNKCYJNY'
        if (not funkcyjny) and (obszar.startswith('FLOT') or 'strategic' in kanal):
            aak_flota[k] += 1
        else:
            aak_detal[k] += 1
    aak_cnt = {k: aak_detal.get(k, 0) + aak_flota.get(k, 0)
               for k in set(aak_detal) | set(aak_flota)}

    # ---- AEK per handlowiec / obszar / kanal ------------------------
    # Uwzgledniamy tylko obszary handlowe. Rekordy z innych obszarow
    # (np. DETALICZNY_reszta) to techniczne duplikaty w bazie AEK -
    # odfiltrowujemy je i raportujemy w filtered_records_count/_clients.
    # aek_licz_reszta=True scala pozostale obszary DETALICZNY_* do detalu
    # (kanal 7.others). Domyslnie False - w bazie AEK siedza tam techniczne
    # duplikaty ubytkow z poprzednich miesiecy.
    licz_reszta = bool((man or {}).get('aek_licz_reszta'))
    OBSZARY = ['DETALICZNY_RD', 'FLOTOWY']
    stan = OrderedDict((o, {'p': 0, 'u': 0, 'd': 0}) for o in OBSZARY)
    aek = defaultdict(lambda: {'p': 0, 'u': 0, 'd': 0, 'p_rd': 0, 'u_rd': 0, 'p_fl': 0, 'u_fl': 0})
    aek_names = {}
    kanaly = defaultdict(int)
    filtered = defaultdict(int)
    for r in b_aek:
        delta = b_aek.s(r, 'delta').upper()
        plus = delta.startswith('PRZYROST')
        minus = delta.startswith('UBYTEK')
        if not (plus or minus):
            continue
        obszar = b_aek.s(r, 'obszar').upper()
        if licz_reszta and obszar.startswith('DETALICZNY'):
            obszar = 'DETALICZNY_RD'
        if obszar not in stan:
            key = (b_aek.s(r, 'model'), pretty_name(b_aek.s(r, 'odpowiedzialny')),
                   b_aek.s(r, 'kanał_rodzaj'), b_aek.s(r, 'obszar'),
                   'przyrost' if plus else 'ubytek', d2s(b_aek.g(r, 'AEK')))
            filtered[key] += 1
            continue
        stan[obszar]['p' if plus else 'u'] += 1
        if plus:
            kan = b_aek.s(r, 'kanał_rodzaj') or 'nieokreślony'
            kanaly[kan] += 1
        who = b_aek.s(r, 'odpowiedzialny')
        if who:
            k = norm(who)
            aek_names.setdefault(k, pretty_name(who))
            e = aek[k]
            e['p' if plus else 'u'] += 1
            # +RD = kanal retail, +Flota = kanaly flotowe (small/medium/strategic fleet)
            sfx = '_rd' if 'retail' in b_aek.s(r, 'kanał_rodzaj').lower() else '_fl'
            e[('p' if plus else 'u') + sfx] += 1
    for v in stan.values():
        v['d'] = v['p'] - v['u']
    for v in aek.values():
        v['d'] = v['p'] - v['u']
    new['stan_aek'] = dict(stan)
    new['kanaly'] = [{'kanal': k, 'aek': v} for k, v in sorted(kanaly.items())]
    new['filtered_records_count'] = sum(filtered.values())
    new['filtered_clients'] = [
        '%d× %s (%s, kanał %s, obszar %s, %s%s) — odfiltrowany z per-handlowiec i stanu obszarów'
        % (n, k[0] or '?', k[1] or 'brak handlowca', k[2] or '?', k[3] or '?', k[4],
           ' z ' + k[5] if k[5] else '')
        for k, n in sorted(filtered.items(), key=lambda x: -x[1])]

    # ---- konfiguracja zespolu (z poprzedniego DATA) -----------------
    old_cele = (old.get('cele') or {}).get('cele') or []
    flota_names = [h.get('name') for h in (old.get('handlowcy_flota') or [])]
    flota_keys = set(norm(n) for n in flota_names)
    detal_prev = [h.get('name') for h in (old.get('handlowcy_detal') or [])]

    roster = []          # [(key, name, short, cel)]
    seen = set()
    for c in old_cele:
        k = norm(c.get('name'))
        if k in seen:
            continue
        seen.add(k)
        roster.append((k, c.get('name'), c.get('short') or (c.get('name') or '').split()[0], i(c.get('cel'))))
    for n in detal_prev + flota_names:
        k = norm(n)
        if k and k not in seen:
            seen.add(k)
            roster.append((k, n, (n or '').split()[0], 0))
    # nowe osoby pojawiajace sie w raporcie
    for k, n in list(aak_names.items()) + list(aek_names.items()):
        if k not in seen:
            seen.add(k)
            roster.append((k, n, n.split()[0], 0))

    # ---- cele (TYLKO DETAL) -----------------------------------------
    cel_zesp = i((old.get('cele') or {}).get('cel_zespolowy'), 0)
    wyk_real = new['realizacja']['detal']['aak']      # AAK detaliczne z REALIZACJA_detal
    cele_rows = []
    for k, name, short, cel in roster:
        if k in flota_keys:
            continue
        wyk = aak_detal.get(k, 0)
        pct = int(round(wyk / cel * 100)) if cel else 0
        cele_rows.append({
            'name': name, 'short': short, 'cel': cel, 'wykonanie': wyk,
            'flota': 0, 'aak_flotowe': aak_flota.get(k, 0), 'pct': pct,
            'brakuje': max(0, cel - wyk),
            'status': 'good' if (cel and pct >= 100) else ('mid' if (cel and pct >= 50) else ('risk' if cel else 'flota')),
        })
    for k, name, short, cel in roster:
        if k not in flota_keys:
            continue
        cele_rows.append({
            'name': name, 'short': short, 'cel': 0,
            'wykonanie': aak_flota.get(k, 0) + aak_detal.get(k, 0),
            'flota': 1, 'aak_flotowe': aak_flota.get(k, 0),
            'pct': 0, 'brakuje': 0, 'status': 'flota',
        })
    cele_rows = [c for c in cele_rows if c['cel'] or c['wykonanie'] or c['flota']]
    cele_rows.sort(key=lambda c: (c['status'] == 'flota', -c['cel'], -c['wykonanie']))

    new['cele'] = dict(old.get('cele') or {})
    new['cele'].update({
        'cel_zespolowy': cel_zesp,
        'wykonanie_realne': wyk_real,
        'brakuje': max(0, cel_zesp - wyk_real),
        'pct': int(round(wyk_real / cel_zesp * 100)) if cel_zesp else 0,
        'uplyw': uplyw,
        'abk_bm': new['realizacja']['detal']['abk_bm'],
        'fk': new['realizacja']['detal']['fk'],
        'cele': cele_rows,
    })

    # ---- handlowcy detal / flota -----------------------------------
    def aek_row(k, name):
        e = aek.get(k) or {'p': 0, 'u': 0, 'd': 0, 'p_rd': 0, 'u_rd': 0, 'p_fl': 0, 'u_fl': 0}
        return {'name': name, 'p': e['p'], 'u': e['u'], 'd': e['d'],
                'p_rd': e['p_rd'], 'u_rd': e['u_rd'], 'p_fl': e['p_fl'], 'u_fl': e['u_fl']}

    hd, hf = [], []
    for k, name, short, cel in roster:
        row = aek_row(k, name)
        if k in flota_keys:
            hf.append(row)
        elif row['p'] or row['u'] or cel:
            hd.append(row)
    hd.sort(key=lambda r: -r['d'])
    hf.sort(key=lambda r: -r['d'])
    new['handlowcy_detal'] = hd
    new['handlowcy_flota'] = hf

    # ---- ranking combined ------------------------------------------
    rank = []
    for k, name, short, cel in roster:
        # detal liczy AAK detaliczne, zespol flotowy - swoje AAK flotowe
        a = aak_flota.get(k, 0) if k in flota_keys else aak_detal.get(k, 0)
        e = (aek.get(k) or {}).get('d', 0)
        if a or e:
            rank.append({'name': name, 'aak': a, 'aek': e, 'suma': a + e})
    rank.sort(key=lambda r: (-r['suma'], -r['aak'], r['name']))
    new['ranking_combined'] = rank

    # ---- magazyn ----------------------------------------------------
    klienc_m, wolne_m = defaultdict(int), defaultdict(int)
    kl = wo = 0
    for r in b_mag:
        model = b_mag.s(r, 'model') or '?'
        if b_mag.s(r, 'rodzaj_zamówienia').upper().endswith('CU'):
            kl += 1
            klienc_m[model] += 1
        else:
            wo += 1
            wolne_m[model] += 1
    top = lambda d: [{'model': m, 'n': n} for m, n in sorted(d.items(), key=lambda x: (-x[1], x[0]))[:8]]
    new['magazyn'] = {'razem': kl + wo, 'klienc': kl, 'wolne': wo,
                      'klienc_modele': top(klienc_m), 'wolne_modele': top(wolne_m)}

    # ---- przeterminowane -------------------------------------------
    def przet_row(sh, r):
        return {
            'vin': sh.s(r, 'vin'),
            'model': sh.s(r, 'model'),
            'pz': d2s(sh.g(r, 'PZ_VGP')),
            'dni': i(sh.g(r, 'AGE_dni')),
            'wiek': i(sh.g(r, 'AGEbm')),
            'typ': 'CU' if sh.s(r, 'rodzaj_zamówienia').upper().endswith('CU')
                   else ('DE' if sh.s(r, 'rodzaj_zamówienia').upper().endswith('DE') else 'FREE'),
            'aak6': pretty_name(sh.s(r, 'sprzedawca')),
        }

    w120, keys120 = [], set()
    for r in b_120:
        w120.append(przet_row(b_120, r))
        keys120.add(b_120.s(r, 'kt') or b_120.s(r, 'vin'))
    w90 = []
    for r in b_90:
        if (b_90.s(r, 'kt') or b_90.s(r, 'vin')) in keys120:
            continue
        w90.append(przet_row(b_90, r))
    w120.sort(key=lambda x: -x['dni'])
    w90.sort(key=lambda x: -x['dni'])
    new['przet'] = {'w120': w120, 'w90': w90, 'w60': []}

    # ---- dostawy (ostatnie 7 dni) ----------------------------------
    cutoff = (rap_d - dt.timedelta(days=7)) if rap_d else None

    def dost_row(sh, r, datecol):
        return {
            'data': d2s(sh.g(r, datecol)),
            'model': sh.s(r, 'model'),
            'vin': sh.s(r, 'vin'),
            'typ': 'CU' if sh.s(r, 'rodzaj_zamówienia').upper().endswith('CU')
                   else ('DE' if sh.s(r, 'rodzaj_zamówienia').upper().endswith('DE') else 'FREE'),
            'sprzedawca': pretty_name(sh.s(r, 'sprzedawca')),
        }

    vgp, aah = [], []
    for r in b_wd:
        d = d2d(b_wd.g(r, 'PZ_VGP'))
        if d and cutoff and d > cutoff:
            vgp.append(dost_row(b_wd, r, 'PZ_VGP'))
    for r in b_mag:
        d = d2d(b_mag.g(r, 'AAH'))
        if d and cutoff and d > cutoff:
            aah.append(dost_row(b_mag, r, 'AAH'))
    vgp.sort(key=lambda x: x['data'], reverse=True)
    aah.sort(key=lambda x: x['data'], reverse=True)
    new['dostawy'] = {'cutoff': d2s(cutoff) if cutoff else '', 'vgp': vgp, 'aah': aah}

    # ---- prognoza ETA (PIPE + MAG_VGP, bez MAG_DE) -----------------
    pmc = defaultdict(lambda: {'cu': 0, 'de': 0})
    per_h = defaultdict(lambda: defaultdict(int))
    for r in b_pl:
        if b_pl.s(r, 'MAG_DE'):
            continue
        eta = d2d(b_pl.g(r, 'ETA'))
        key = eta.strftime('%Y-%m') if eta else 'brak'
        typ = 'cu' if b_pl.s(r, 'rodzaj_zamówienia').upper().endswith('CU') else 'de'
        pmc[key][typ] += 1
        who = b_pl.s(r, 'sprzedawca')
        if who:
            per_h[norm(who)][key] += 1
            aak_names.setdefault(norm(who), pretty_name(who))
    mies = sorted([k for k in pmc if k != 'brak'])
    order = mies + (['brak'] if 'brak' in pmc else [])
    new['prognoza_mies'] = order
    new['prognoza'] = [{'mc': k, 'cu': pmc[k]['cu'], 'de': pmc[k]['de'],
                        'razem': pmc[k]['cu'] + pmc[k]['de']} for k in order]

    name_of = {}
    for k, name, short, cel in roster:
        name_of[k] = name
    ph = []
    for k, cnts in per_h.items():
        row = {'name': name_of.get(k) or aak_names.get(k) or aek_names.get(k) or pretty_name(k)}
        tot = 0
        for m in mies:
            row[m] = cnts.get(m, 0)
            tot += row[m]
        row['brak'] = cnts.get('brak', 0)
        row['razem'] = tot + row['brak']
        ph.append(row)
    ph.sort(key=lambda r: -r['razem'])
    new['prognoza_handlowcy'] = ph

    # ---- dane wprowadzane recznie (manual.json) ---------------------
    if man:
        apply_manual(new, man, rap_d, roster, flota_keys)

    return new, {
        'data_raportu': data_rap,
        'aak': new['realizacja']['razem']['aak'],
        'aek_d': sum(v['d'] for v in stan.values()),
        'mag': new['magazyn']['razem'],
        'p120': len(w120), 'p90': len(w90),
        'handlowcy': len(cele_rows),
    }


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--html', required=True)
    ap.add_argument('--out')
    ap.add_argument('--manual', help='JSON z danymi recznymi / z arkuszy Google')
    ap.add_argument('--dump-json')
    a = ap.parse_args()

    html = open(a.html, encoding='utf-8').read()
    decl, s, e = find_data_span(html)
    old = json.loads(html[s:e])
    man = json.load(open(a.manual, encoding='utf-8')) if a.manual else None

    new, summary = build(a.xlsx, old, man)

    body = json.dumps(new, ensure_ascii=False, separators=(', ', ': '))
    out_html = html[:s] + body + html[e:]

    dest = a.out or a.html
    open(dest, 'w', encoding='utf-8', newline='\n').write(out_html)
    if a.dump_json:
        open(a.dump_json, 'w', encoding='utf-8').write(json.dumps(new, ensure_ascii=False, indent=1))

    # walidacja: czy nowy DATA da sie odczytac
    chk = open(dest, encoding='utf-8').read()
    d2, s2, e2 = find_data_span(chk)
    json.loads(chk[s2:e2])

    print('OK  %s' % dest)
    for k, v in summary.items():
        print('    %-14s %s' % (k, v))


if __name__ == '__main__':
    main()
